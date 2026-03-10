"""Comprehensive tests for io.py file loaders.

Tests focus on XLSX/CSV loading with column resolution, sheet detection, and edge cases.
"""

import sys
import os
import tempfile
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import pandas as pd
import openpyxl
from llm_detector.io import load_xlsx, load_csv, _col_letter_to_index

PASSED = 0
FAILED = 0


def check(label, condition, detail=""):
    global PASSED, FAILED
    if condition:
        PASSED += 1
        print(f"  [PASS] {label}")
    else:
        FAILED += 1
        print(f"  [FAIL] {label}  -- {detail}")


def test_col_letter_to_index():
    """Test column letter/number to index conversion."""
    print("\n-- COLUMN LETTER TO INDEX --")

    # Single letters A-Z
    check("A -> 0", _col_letter_to_index('A') == 0)
    check("B -> 1", _col_letter_to_index('B') == 1)
    check("Z -> 25", _col_letter_to_index('Z') == 25)
    check("lowercase a -> 0", _col_letter_to_index('a') == 0)

    # 1-based numbers
    check("'1' -> 0", _col_letter_to_index('1') == 0)
    check("'2' -> 1", _col_letter_to_index('2') == 1)
    check("'10' -> 9", _col_letter_to_index('10') == 9)

    # Invalid inputs return None
    check("'invalid' -> None", _col_letter_to_index('invalid') is None)
    check("'AA' -> None", _col_letter_to_index('AA') is None)
    check("'0' -> None", _col_letter_to_index('0') is None)
    check("'-1' -> None", _col_letter_to_index('-1') is None)


def test_load_xlsx_basic():
    """Test basic XLSX loading with default sheet."""
    print("\n-- LOAD XLSX BASIC --")

    # Create a temporary XLSX file
    with tempfile.NamedTemporaryFile(mode='wb', suffix='.xlsx', delete=False) as f:
        filepath = f.name

    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        # Header row
        ws.append(['task_id', 'prompt', 'occupation', 'attempter_name', 'pipeline_stage_name'])

        # Data rows
        ws.append(['task_001', 'This is a test prompt with more than fifty characters in it to pass the length check', 'doctor', 'alice', 'stage1'])
        ws.append(['task_002', 'Another test prompt that is long enough to be included in the results from the loader', 'engineer', 'bob', 'stage2'])
        ws.append(['task_003', 'Short', 'teacher', 'charlie', 'stage3'])  # Too short

        wb.save(filepath)
        wb.close()

        tasks = load_xlsx(filepath)
        check("Load XLSX returns 2 tasks (1 filtered out)", len(tasks) == 2)
        check("First task has correct task_id", tasks[0]['task_id'] == 'task_001')
        check("First task has correct prompt", 'test prompt' in tasks[0]['prompt'])
        check("First task has correct occupation", tasks[0]['occupation'] == 'doctor')
        check("First task has correct attempter", tasks[0]['attempter'] == 'alice')
        check("First task has correct stage", tasks[0]['stage'] == 'stage1')
        check("Second task has correct task_id", tasks[1]['task_id'] == 'task_002')

    finally:
        os.unlink(filepath)


def test_load_xlsx_default_sheet_detection():
    """Test default sheet detection (FullTaskX, etc.)."""
    print("\n-- LOAD XLSX DEFAULT SHEET DETECTION --")

    with tempfile.NamedTemporaryFile(mode='wb', suffix='.xlsx', delete=False) as f:
        filepath = f.name

    try:
        wb = openpyxl.Workbook()

        # Create a decoy first sheet
        ws1 = wb.active
        ws1.title = "Summary"
        ws1.append(['header'])
        ws1.append(['dummy data'])

        # Create FullTaskX sheet (should be auto-selected)
        ws2 = wb.create_sheet("FullTaskX")
        ws2.append(['prompt', 'task_id'])
        ws2.append(['This is a test prompt with more than fifty characters in it to pass the filter', 'task_001'])

        wb.save(filepath)
        wb.close()

        tasks = load_xlsx(filepath)
        check("FullTaskX sheet auto-detected", len(tasks) == 1)
        check("Task loaded from FullTaskX", tasks[0]['task_id'] == 'task_001')

    finally:
        os.unlink(filepath)


def test_load_xlsx_positional_columns():
    """Test positional column references (A, B, C or 1, 2, 3)."""
    print("\n-- LOAD XLSX POSITIONAL COLUMNS --")

    with tempfile.NamedTemporaryFile(mode='wb', suffix='.xlsx', delete=False) as f:
        filepath = f.name

    try:
        wb = openpyxl.Workbook()
        ws = wb.active

        # Header row with weird names
        ws.append(['weird_col_1', 'weird_col_2', 'weird_col_3'])
        ws.append(['task_001', 'This is a test prompt with more than fifty characters in it to pass the length check', 'doctor'])

        wb.save(filepath)
        wb.close()

        # Use positional letter references
        tasks = load_xlsx(filepath, prompt_col='B', id_col='A', occ_col='C')
        check("Positional letters work", len(tasks) == 1)
        check("Column B loaded as prompt", 'test prompt' in tasks[0]['prompt'])
        check("Column A loaded as task_id", tasks[0]['task_id'] == 'task_001')
        check("Column C loaded as occupation", tasks[0]['occupation'] == 'doctor')

    finally:
        os.unlink(filepath)


def test_load_xlsx_positional_numbers():
    """Test positional column references using 1-based numbers."""
    print("\n-- LOAD XLSX POSITIONAL NUMBERS --")

    with tempfile.NamedTemporaryFile(mode='wb', suffix='.xlsx', delete=False) as f:
        filepath = f.name

    try:
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.append(['col1', 'col2', 'col3'])
        ws.append(['task_001', 'This is a test prompt with more than fifty characters in it to pass the length check', 'doctor'])

        wb.save(filepath)
        wb.close()

        # Use 1-based number strings
        tasks = load_xlsx(filepath, prompt_col='2', id_col='1', occ_col='3')
        check("Positional numbers work", len(tasks) == 1)
        check("Column 2 loaded as prompt", 'test prompt' in tasks[0]['prompt'])
        check("Column 1 loaded as task_id", tasks[0]['task_id'] == 'task_001')
        check("Column 3 loaded as occupation", tasks[0]['occupation'] == 'doctor')

    finally:
        os.unlink(filepath)


def test_load_xlsx_fuzzy_column_matching():
    """Test fuzzy/substring column matching."""
    print("\n-- LOAD XLSX FUZZY MATCHING --")

    with tempfile.NamedTemporaryFile(mode='wb', suffix='.xlsx', delete=False) as f:
        filepath = f.name

    try:
        wb = openpyxl.Workbook()
        ws = wb.active

        # Headers with variations
        ws.append(['my_task_id_field', 'user_prompt_text', 'user_occupation'])
        ws.append(['task_001', 'This is a test prompt with more than fifty characters in it to pass the check', 'doctor'])

        wb.save(filepath)
        wb.close()

        # Fuzzy matching should find 'task_id' in 'my_task_id_field', 'prompt' in 'user_prompt_text', etc.
        tasks = load_xlsx(filepath)
        check("Fuzzy matching finds columns", len(tasks) == 1)
        check("Fuzzy match finds prompt", 'test prompt' in tasks[0]['prompt'])
        check("Fuzzy match finds task_id", tasks[0]['task_id'] == 'task_001')
        check("Fuzzy match finds occupation", tasks[0]['occupation'] == 'doctor')

    finally:
        os.unlink(filepath)


def test_load_xlsx_short_prompt_filtering():
    """Test that prompts < 50 chars are filtered out."""
    print("\n-- LOAD XLSX SHORT PROMPT FILTERING --")

    with tempfile.NamedTemporaryFile(mode='wb', suffix='.xlsx', delete=False) as f:
        filepath = f.name

    try:
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.append(['prompt', 'task_id'])
        ws.append(['Short prompt', 'task_001'])  # < 50 chars
        ws.append(['This is a much longer prompt that exceeds fifty characters and should be included', 'task_002'])
        ws.append(['Also short', 'task_003'])  # < 50 chars

        wb.save(filepath)
        wb.close()

        tasks = load_xlsx(filepath)
        check("Only long prompts included", len(tasks) == 1)
        check("Correct task retained", tasks[0]['task_id'] == 'task_002')

    finally:
        os.unlink(filepath)


def test_load_xlsx_empty_workbook():
    """Test handling of empty workbook."""
    print("\n-- LOAD XLSX EMPTY WORKBOOK --")

    with tempfile.NamedTemporaryFile(mode='wb', suffix='.xlsx', delete=False) as f:
        filepath = f.name

    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        # No data

        wb.save(filepath)
        wb.close()

        tasks = load_xlsx(filepath)
        check("Empty workbook returns empty list", len(tasks) == 0)

    finally:
        os.unlink(filepath)


def test_load_csv_basic():
    """Test basic CSV loading."""
    print("\n-- LOAD CSV BASIC --")

    with tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False, newline='') as f:
        filepath = f.name
        f.write('task_id,prompt,occupation,attempter_name,pipeline_stage_name\n')
        f.write('task_001,"This is a test prompt with more than fifty characters in it to pass the length check",doctor,alice,stage1\n')
        f.write('task_002,"Another test prompt that is long enough to be included in the results from loader",engineer,bob,stage2\n')

    try:
        tasks = load_csv(filepath)
        check("Load CSV returns 2 tasks", len(tasks) == 2)
        check("First task has correct task_id", tasks[0]['task_id'] == 'task_001')
        check("First task has correct prompt", 'test prompt' in tasks[0]['prompt'])
        check("First task has correct occupation", tasks[0]['occupation'] == 'doctor')
        check("First task has correct attempter", tasks[0]['attempter'] == 'alice')
        check("First task has correct stage", tasks[0]['stage'] == 'stage1')

    finally:
        os.unlink(filepath)


def test_load_csv_positional_columns():
    """Test CSV with positional column references."""
    print("\n-- LOAD CSV POSITIONAL COLUMNS --")

    with tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False, newline='') as f:
        filepath = f.name
        f.write('col1,col2,col3\n')
        f.write('task_001,"This is a test prompt with more than fifty characters in it to pass the check",doctor\n')

    try:
        # Use positional letter references
        tasks = load_csv(filepath, prompt_col='B', id_col='A', occ_col='C')
        check("CSV positional letters work", len(tasks) == 1)
        check("Column B loaded as prompt", 'test prompt' in tasks[0]['prompt'])
        check("Column A loaded as task_id", tasks[0]['task_id'] == 'task_001')
        check("Column C loaded as occupation", tasks[0]['occupation'] == 'doctor')

    finally:
        os.unlink(filepath)


def test_load_csv_fuzzy_matching():
    """Test CSV with fuzzy column matching."""
    print("\n-- LOAD CSV FUZZY MATCHING --")

    with tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False, newline='') as f:
        filepath = f.name
        f.write('my_task_id_field,user_prompt_text,user_occupation\n')
        f.write('task_001,"This is a test prompt with more than fifty characters in it to pass check",doctor\n')

    try:
        tasks = load_csv(filepath)
        check("CSV fuzzy matching finds columns", len(tasks) == 1)
        check("CSV fuzzy match finds prompt", 'test prompt' in tasks[0]['prompt'])
        check("CSV fuzzy match finds task_id", tasks[0]['task_id'] == 'task_001')

    finally:
        os.unlink(filepath)


def test_load_csv_short_prompt_filtering():
    """Test CSV short prompt filtering."""
    print("\n-- LOAD CSV SHORT PROMPT FILTERING --")

    with tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False, newline='') as f:
        filepath = f.name
        f.write('prompt,task_id\n')
        f.write('Short,task_001\n')
        f.write('"This is a much longer prompt that exceeds fifty characters and should be included",task_002\n')
        f.write('Also short,task_003\n')

    try:
        tasks = load_csv(filepath)
        check("CSV filters short prompts", len(tasks) == 1)
        check("CSV correct task retained", tasks[0]['task_id'] == 'task_002')

    finally:
        os.unlink(filepath)


def test_load_csv_missing_prompt_column():
    """Test CSV error handling when prompt column not found."""
    print("\n-- LOAD CSV MISSING PROMPT COLUMN --")

    with tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False, newline='') as f:
        filepath = f.name
        f.write('task_id,description\n')
        f.write('task_001,"Some description that is not a prompt column"\n')

    try:
        # Capture output to check for error message
        import io
        from contextlib import redirect_stdout

        output = io.StringIO()
        with redirect_stdout(output):
            tasks = load_csv(filepath)

        check("CSV missing prompt returns empty", len(tasks) == 0)
        check("CSV missing prompt prints error", 'ERROR' in output.getvalue())

    finally:
        os.unlink(filepath)


def test_load_xlsx_missing_prompt_column():
    """Test XLSX error handling when prompt column not found."""
    print("\n-- LOAD XLSX MISSING PROMPT COLUMN --")

    with tempfile.NamedTemporaryFile(mode='wb', suffix='.xlsx', delete=False) as f:
        filepath = f.name

    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['task_id', 'description'])
        ws.append(['task_001', 'Some description'])

        wb.save(filepath)
        wb.close()

        import io
        from contextlib import redirect_stdout

        output = io.StringIO()
        with redirect_stdout(output):
            tasks = load_xlsx(filepath)

        check("XLSX missing prompt returns empty", len(tasks) == 0)
        check("XLSX missing prompt prints error", 'ERROR' in output.getvalue())

    finally:
        os.unlink(filepath)


if __name__ == '__main__':
    test_col_letter_to_index()
    test_load_xlsx_basic()
    test_load_xlsx_default_sheet_detection()
    test_load_xlsx_positional_columns()
    test_load_xlsx_positional_numbers()
    test_load_xlsx_fuzzy_column_matching()
    test_load_xlsx_short_prompt_filtering()
    test_load_xlsx_empty_workbook()
    test_load_csv_basic()
    test_load_csv_positional_columns()
    test_load_csv_fuzzy_matching()
    test_load_csv_short_prompt_filtering()
    test_load_csv_missing_prompt_column()
    test_load_xlsx_missing_prompt_column()

    print(f"\n{'='*60}")
    print(f"TOTAL: {PASSED} passed, {FAILED} failed")
    print(f"{'='*60}")

    if FAILED > 0:
        sys.exit(1)
